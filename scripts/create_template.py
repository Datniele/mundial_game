"""
Genera il file Excel template da distribuire ai partecipanti.
Uso: python scripts/create_template.py [output_path]
"""
import sys
import json
from pathlib import Path
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Protection
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).parent.parent
FIXTURES_PATH = ROOT / "data" / "fixtures" / "fixtures.json"
OUTPUT_PATH = ROOT / "template_pronostici.xlsx"

YELLOW = PatternFill("solid", fgColor="FFFF99")
GREEN = PatternFill("solid", fgColor="C6EFCE")
BLUE = PatternFill("solid", fgColor="BDD7EE")
GRAY = PatternFill("solid", fgColor="D9D9D9")
ORANGE = PatternFill("solid", fgColor="FCE4D6")

BOLD = Font(bold=True)
HEADER_FONT = Font(bold=True, color="FFFFFF")
CENTER = Alignment(horizontal="center", vertical="center")


def _header_row(ws, headers: list[str], fill: PatternFill, row: int = 1):
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=h)
        cell.fill = fill
        cell.font = HEADER_FONT
        cell.alignment = CENTER


def _set_col_widths(ws, widths: list[int]):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def build_istruzioni(wb: openpyxl.Workbook):
    ws = wb.create_sheet("ISTRUZIONI", 0)
    ws.sheet_view.showGridLines = False
    ws["A1"] = "ISTRUZIONI PER LA COMPILAZIONE"
    ws["A1"].font = Font(bold=True, size=14)
    istruzioni = [
        "",
        "PUOI CONSEGNARE I PRONOSTICI IN PIU' RIPRESE, FASE PER FASE.",
        "Ogni caricamento aggiorna solo le fasi presenti nel file; le altre restano invariate.",
        "",
        "1. Compila i fogli delle fasi che vuoi inviare (non obbligatorio compilarli tutti subito).",
        "2. Foglio GIRONI: inserisci i gol previsti per ogni partita (colonne gialle).",
        "   Il torneo ha 12 gironi (A-L) con 4 squadre ciascuno, per un totale di 72 partite.",
        "3. Foglio CLASSIFICHE_GIRONI: inserisci chi finirà 1° e 2° in ogni girone.",
        "4. Fogli SEDICESIMI → FINALE: inserisci le squadre che prevedi si qualifichino e il risultato.",
        "   Avanzano ai sedicesimi: i primi 2 di ogni girone (24 squadre) + le 8 migliori terze (32 totali).",
        "5. Non modificare i nomi dei fogli né le intestazioni delle colonne.",
        "6. Salva il file con il nome: nome_cognome.xlsx (es. mario_rossi.xlsx).",
        "7. Consegna il file al referente. Puoi riconsegnarlo più volte per aggiungere fasi successive.",
        "",
        "SISTEMA DI PUNTEGGIO (modificabile nella sezione Configurazione dell'app):",
        "  • Risultato esatto girone:          3 punti",
        "  • Esito corretto (V/P/N) girone:    1 punto",
        "  • Classifica girone identica:       5 punti",
        "  • Classifica girone parziale:       2 punti",
        "  • Chi passa il turno (eliminaz.):   2 punti",
        "  • Risultato esatto (eliminaz.):     4 punti",
    ]
    for i, text in enumerate(istruzioni, 2):
        ws[f"A{i}"] = text
    ws.column_dimensions["A"].width = 70


def build_gironi(wb: openpyxl.Workbook, fixtures: dict):
    ws = wb.create_sheet("GIRONI")
    headers = ["ID Partita", "Girone", "Squadra Casa", "Squadra Ospite", "Gol Casa", "Gol Ospite"]
    _header_row(ws, headers, BLUE)
    _set_col_widths(ws, [12, 8, 20, 20, 10, 10])

    for r, match in enumerate(fixtures["group_matches"], 2):
        ws.cell(r, 1, match["id"]).fill = GRAY
        ws.cell(r, 2, match["group"]).alignment = CENTER
        ws.cell(r, 3, match["home"])
        ws.cell(r, 4, match["away"])
        for col in (5, 6):
            c = ws.cell(r, col, "")
            c.fill = YELLOW
            c.alignment = CENTER

    ws.protection.sheet = False


def build_classifiche(wb: openpyxl.Workbook, fixtures: dict):
    ws = wb.create_sheet("CLASSIFICHE_GIRONI")
    headers = ["Girone", "1° Posto", "2° Posto"]
    _header_row(ws, headers, GREEN)
    _set_col_widths(ws, [8, 22, 22])

    for r, group in enumerate(sorted(fixtures["groups"].keys()), 2):
        ws.cell(r, 1, group).alignment = CENTER
        for col in (2, 3):
            c = ws.cell(r, col, "")
            c.fill = YELLOW


def build_knockout_sheet(wb: openpyxl.Workbook, sheet_name: str, n_matches: int, prefix: str):
    ws = wb.create_sheet(sheet_name)
    headers = ["ID Partita", "Squadra 1", "Squadra 2", "Gol Sq1", "Gol Sq2"]
    _header_row(ws, headers, ORANGE)
    _set_col_widths(ws, [12, 22, 22, 10, 10])

    for i in range(1, n_matches + 1):
        r = i + 1
        ws.cell(r, 1, f"{prefix}{i}").fill = GRAY
        for col in (2, 3, 4, 5):
            c = ws.cell(r, col, "")
            c.fill = YELLOW
            c.alignment = CENTER


def main():
    out = Path(sys.argv[1]) if len(sys.argv) > 1 else OUTPUT_PATH

    with open(FIXTURES_PATH, encoding="utf-8") as f:
        fixtures = json.load(f)

    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # rimuove foglio di default

    build_istruzioni(wb)
    build_gironi(wb, fixtures)
    build_classifiche(wb, fixtures)

    knockout_config = [
        ("SEDICESIMI", 16, "S"),
        ("OTTAVI", 8, "O"),
        ("QUARTI", 4, "Q"),
        ("SEMIFINALI", 2, "SF"),
        ("3_POSTO", 1, "3P"),
        ("FINALE", 1, "F"),
    ]
    for sheet_name, n, prefix in knockout_config:
        build_knockout_sheet(wb, sheet_name, n, prefix)

    wb.save(out)
    print(f"Template salvato in: {out}")


if __name__ == "__main__":
    main()
